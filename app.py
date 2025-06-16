import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Earned Commission Table Extractor", layout="wide")
st.title("⚙️ Earned Commission Table Extractor (Automated)")

# --- Configuration Constants for the Specific File ---
AUTO_FILL_START_ROW = 23
AUTO_FILL_END_ROW = 36
AUTO_FILL_START_COL = 2
AUTO_FILL_END_COL = 5

ROWS_TO_AUTO_REMOVE = [8, 10, 13]
ROWS_TO_AUTO_COMBINE = [11, 12]
AUTO_COMBINED_ROW_NAME = "MT - Without FV"

# --- Utility Functions (Cached) ---
# @st.cache_resource is good for workbook loading as it's a heavy object
@st.cache_resource(ttl=3600)
def load_workbook_from_bytesio(file_buffer):
    """Loads an OpenPyXL workbook from a BytesIO object."""
    file_buffer.seek(0)
    return openpyxl.load_workbook(file_buffer, data_only=True)

# @st.cache_data is good for DataFrame creation from static inputs
@st.cache_data(ttl=3600)
def get_initial_dataframe(_workbook, sheet_name, start_row, end_row, start_col, end_col, use_first_row_as_header):
    """
    Extracts a DataFrame from a specified range within an OpenPyXL worksheet.
    Handles header logic and ensures unique column names.
    """
    ws = _workbook[sheet_name]

    start_row = max(1, start_row)
    end_row = max(start_row, end_row)
    start_col = max(1, start_col)
    end_col = max(start_col, end_col)

    data = []
    try:
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
            data.append(list(row))
    except Exception as e:
        st.error(f"Error reading specified range from sheet: {e}. Please check your row/column inputs.")
        return pd.DataFrame()

    if not data:
        return pd.DataFrame()

    if use_first_row_as_header and len(data) > 0:
        raw_headers = list(data[0])
        rows = data[1:]
    else:
        raw_headers = [f"Column_{i+1}" for i in range(end_col - start_col + 1)]
        rows = data

    headers = []
    seen = {}
    for h in raw_headers:
        h_str = str(h) if h is not None and str(h).strip() != "" else "Unnamed"
        if h_str in seen:
            seen[h_str] += 1
            h_str = f"{h_str}_{seen[h_str]}"
        else:
            seen[h_str] = 0
        headers.append(h_str)
    
    adjusted_rows = []
    expected_cols = len(headers)
    for row in rows:
        if len(row) < expected_cols:
            adjusted_rows.append(list(row) + [None] * (expected_cols - len(row)))
        else:
            adjusted_rows.append(list(row[:expected_cols]))
            
    df_result = pd.DataFrame(adjusted_rows, columns=headers)
    df_result = df_result.dropna(how="all")

    if 'Order' not in df_result.columns:
        df_result.insert(0, 'Order', range(1, len(df_result) + 1))

    return df_result

# Function to convert DataFrame to Excel for download
def to_excel_specific(df_to_save):
    """Converts a DataFrame to an Excel file in BytesIO object."""
    output = BytesIO()
    if not df_to_save.empty:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_to_save.to_excel(writer, index=False, sheet_name="ModifiedTable")
    output.seek(0)
    return output

# --- Main App Logic ---
with st.sidebar:
    st.header("Upload Excel File")
    # Using a dynamic key for file_uploader ensures it truly resets if the user
    # "re-uploads" the same file, by forcing Streamlit to treat it as a new widget instance.
    # However, Streamlit's file_id is usually sufficient for actual file content changes.
    # We'll rely on file_id for content change detection.
    uploaded_file = st.file_uploader("Upload your specific Excel file", type=["xlsx", "xls"])
    st.markdown("---")

if uploaded_file is not None:
    try:
        # Get a unique identifier for the current file to manage session state
        current_file_id = uploaded_file.file_id

        # Load the workbook and get sheet names immediately
        with st.spinner("Loading Excel file..."):
            wb = load_workbook_from_bytesio(uploaded_file)
        st.success("File loaded successfully!")

        sheet_names = wb.sheetnames
        
        with st.sidebar:
            st.header("Sheet Selection")
            default_sheet_index = 0
            if "Sheet1" in sheet_names: # Prioritize "Sheet1" if it exists
                default_sheet_index = sheet_names.index("Sheet1")
            
            selected_sheet = st.selectbox(
                "Select the relevant sheet", 
                sheet_names, 
                index=default_sheet_index, 
                key="selected_sheet_sidebar_specific" # Unique key
            )
            ws = wb[selected_sheet]
            max_row = ws.max_row
            max_column = ws.max_column
            st.info(f"Sheet dimensions: {max_row} rows, {max_column} columns")
            st.markdown("---")

        st.markdown("### Automatically Process Specific Table")

        # Auto-fill toggle for predefined range (defaults to True)
        auto_fill_toggle_specific = st.toggle(
            f"Use predefined range (Rows {AUTO_FILL_START_ROW}-{AUTO_FILL_END_ROW}, Cols {AUTO_FILL_START_COL}-{AUTO_FILL_END_COL})",
            value=True,
            key="auto_fill_toggle_switch_specific" # Unique key
        )
        
        # Determine current range based on toggle
        if auto_fill_toggle_specific:
            current_start_row = AUTO_FILL_START_ROW
            current_end_row = AUTO_FILL_END_ROW
            current_start_col = AUTO_FILL_START_COL
            current_end_col = AUTO_FILL_END_COL
            current_use_header = True
            st.info("Predefined table range is active.")
        else:
            # Fallback if toggle is somehow turned off, though UI doesn't allow it
            st.warning("Manual adjustment is not available in this simplified app version. Please use the predefined range.")
            current_start_row = AUTO_FILL_START_ROW
            current_end_row = AUTO_FILL_END_ROW
            current_start_col = AUTO_FILL_START_COL
            current_end_col = AUTO_FILL_END_COL
            current_use_header = True

        # Create a unique identifier for the *current state of initial data extraction*
        # This includes file ID, selected sheet, and range parameters.
        current_initial_data_hash = (
            f"{current_file_id}-"
            f"{selected_sheet}-"
            f"{current_start_row}-"
            f"{current_end_row}-"
            f"{current_start_col}-"
            f"{current_end_col}-"
            f"{current_use_header}"
        )

        # Initialize or reset DataFrame and processing flags based on initial data hash
        if "last_initial_data_hash" not in st.session_state or st.session_state.last_initial_data_hash != current_initial_data_hash:
            with st.spinner("Initializing table from file..."):
                df_initial = get_initial_dataframe(wb, selected_sheet,
                                                    current_start_row, current_end_row,
                                                    current_start_col, current_end_col,
                                                    current_use_header)
                st.session_state.current_df_specific = df_initial.copy()
                st.session_state.history_specific = [] # Clear history on new file/selection
                st.session_state.last_initial_data_hash = current_initial_data_hash
                st.session_state.auto_processing_done = False # Reset auto-processing flag

                if not df_initial.empty:
                    st.success("Table initialized. Applying automatic processes...")
                else:
                    st.info("No data found for the predefined range in the selected sheet.")
                
                # Rerun immediately after initializing to proceed with auto-processing
                st.rerun()

        # --- Automatic Row Filtering and Combination (Applied only once per new file/initial data) ---
        if not st.session_state.auto_processing_done and not st.session_state.current_df_specific.empty:
            
            df_to_process = st.session_state.current_df_specific.copy()
            modified = False

            # Automatic Row Filtering
            st.markdown("### 🗑️ Automatic Row Filtering")
            auto_remove_toggle = st.checkbox(
                f"Automatically remove rows with 'Order' numbers: {', '.join(map(str, ROWS_TO_AUTO_REMOVE))}",
                value=True,
                key="auto_remove_rows_toggle_specific" # Unique key
            )

            if auto_remove_toggle and 'Order' in df_to_process.columns:
                original_row_count = len(df_to_process)
                df_temp_remove = df_to_process.copy()
                df_temp_remove['Order_numeric'] = pd.to_numeric(df_temp_remove['Order'], errors='coerce')

                rows_to_keep_mask = ~df_temp_remove['Order_numeric'].isin(ROWS_TO_AUTO_REMOVE)
                
                if not rows_to_keep_mask.all():
                    df_to_process = df_temp_remove[rows_to_keep_mask].drop(columns=['Order_numeric']).reset_index(drop=True)
                    removed_count = original_row_count - len(df_to_process)
                    st.success(f"Automatically removed {removed_count} row(s) based on predefined order numbers.")
                    modified = True
                else:
                    st.info("No rows to remove based on automatic filter settings.")
            st.markdown("---")

            # Automatic Row Combination
            st.markdown("### 🔗 Automatic Row Combination")
            auto_combine_toggle = st.checkbox(
                f"Automatically combine rows with 'Order' numbers: {', '.join(map(str, ROWS_TO_AUTO_COMBINE))} and rename to '{AUTO_COMBINED_ROW_NAME}'",
                value=True,
                key="auto_combine_rows_toggle_specific" # Unique key
            )

            if auto_combine_toggle and 'Order' in df_to_process.columns and len(ROWS_TO_AUTO_COMBINE) > 1:
                df_temp_combine = df_to_process.copy()
                df_temp_combine['Order_numeric_combine'] = pd.to_numeric(df_temp_combine['Order'], errors='coerce')

                indices_to_combine = df_temp_combine[df_temp_combine['Order_numeric_combine'].isin(ROWS_TO_AUTO_COMBINE)].index.tolist()

                if len(indices_to_combine) >= 2:
                    combined_row_data = {}
                    selected_df_for_auto_combine = df_to_process.loc[indices_to_combine]

                    for col_idx, col in enumerate(df_to_process.columns):
                        if pd.api.types.is_numeric_dtype(df_to_process[col]):
                            combined_row_data[col] = selected_df_for_auto_combine[col].sum()
                        else:
                            joined_value = " / ".join(selected_df_for_auto_combine[col].dropna().astype(str).tolist())
                            combined_row_data[col] = joined_value
                        
                    # Handle the 'Order' column and the first non-'Order' column for the new row name
                    if 'Order' in df_to_process.columns:
                        combined_row_data['Order'] = 999999 # Assign a high order number to place it at the end initially
                    
                    first_non_order_col = next((col for col in df_to_process.columns if col != 'Order'), None)
                    if first_non_order_col:
                        combined_row_data[first_non_order_col] = AUTO_COMBINED_ROW_NAME

                    combined_df_new = pd.DataFrame([combined_row_data], columns=df_to_process.columns)
                    
                    remaining_df = df_to_process.drop(index=indices_to_combine).reset_index(drop=True)
                    df_to_process = pd.concat([remaining_df, combined_df_new], ignore_index=True)
                    
                    st.success(f"Automatically combined rows with Order {', '.join(map(str, ROWS_TO_AUTO_COMBINE))} into '{AUTO_COMBINED_ROW_NAME}'.")
                    modified = True
                else:
                    st.warning(f"Could not auto-combine. Found {len(indices_to_combine)} row(s) with order numbers {', '.join(map(str, ROWS_TO_AUTO_COMBINE))}. At least 2 are required to combine.")
            st.markdown("---")

            # Update session state with processed DataFrame and set flag
            if modified:
                st.session_state.current_df_specific = df_to_process.copy()
            st.session_state.auto_processing_done = True
            st.rerun() # Rerun once auto-processing is complete to show the result

        # --- Display and Editing UI ---
        # This part will always display the current_df_specific, which will have
        # either the initial data, or the auto-processed data, or user-edited data.
        if not st.session_state.current_df_specific.empty:
            st.subheader("✏️ Review and Edit Table (Directly in Table)")
            st.info("You can directly edit cells in the table. To reorder rows, edit the numbers in the 'Order' column. To delete a row, click the 'X' button on the right of the row.")

            # Ensure 'Order' column is numeric for proper sorting and data editor
            # Use errors='coerce' to turn non-numeric into NaN, then fillna(0) and convert to int
            st.session_state.current_df_specific['Order'] = pd.to_numeric(st.session_state.current_df_specific['Order'], errors='coerce').fillna(0).astype(int)

            edited_df_specific = st.data_editor(
                st.session_state.current_df_specific,
                num_rows="dynamic", # Allows adding/deleting rows directly in the editor
                use_container_width=True,
                column_config={
                    "Order": st.column_config.NumberColumn(
                        "Order",
                        help="Assign a number to reorder rows. Drag and drop rows within the editor is also possible.",
                        default=0,
                        step=1,
                        format="%d"
                    )
                },
                key="main_data_editor_specific" # Unique key for the data editor
            )

            # Check if edited_df is different from current_df
            # st.data_editor handles internal state, so explicit .equals check and rerun is good.
            if not edited_df_specific.equals(st.session_state.current_df_specific):
                # Only save to history if there's an actual change from user editing
                # This prevents saving the intermediate states of auto-processing to history
                if st.session_state.auto_processing_done: # Ensure auto-processing is done before logging user edits
                    st.session_state.history_specific.append(st.session_state.current_df_specific.copy())
                st.session_state.current_df_specific = edited_df_specific.copy()
                st.success("Changes detected. Apply order or continue editing.")
                st.rerun() # Rerun to ensure the changes are reflected and potentially trigger further UI updates

            if st.button("Apply Manual Row Order Changes", key="apply_order_specific_manual"):
                if 'Order' in st.session_state.current_df_specific.columns:
                    temp_df = st.session_state.current_df_specific.copy()
                    # Handle duplicate order numbers gracefully for sorting
                    temp_df['Order_temp_sort'] = temp_df['Order']
                    if temp_df['Order_temp_sort'].duplicated().any():
                        # Append a unique identifier for stable sort among duplicates
                        temp_df['Order_temp_sort'] = temp_df['Order'].astype(str) + '.' + temp_df.groupby('Order_temp_sort').cumcount().astype(str)
                        temp_df['Order_temp_sort'] = pd.to_numeric(temp_df['Order_temp_sort'], errors='coerce') # Convert back to numeric for sorting
                        
                    st.session_state.current_df_specific = temp_df.sort_values(by='Order_temp_sort', ascending=True).drop(columns=['Order_temp_sort']).reset_index(drop=True)
                    st.success("Rows reordered successfully!")
                    st.rerun() # Rerun to display the sorted table
                else:
                    st.warning("No 'Order' column found to reorder rows.")
            
            st.subheader("📋 Final Processed Table")
            # Display the final, non-all-NA rows of the table
            final_df_specific = st.session_state.current_df_specific.dropna(how="all").reset_index(drop=True)
            # Remove 'Order' column for final display and download
            if 'Order' in final_df_specific.columns:
                final_df_specific = final_df_specific.drop(columns=['Order'])
            
            st.dataframe(final_df_specific, use_container_width=True)

            st.subheader("📥 Download Modified Table")
            excel_data_specific = to_excel_specific(final_df_specific)
            st.download_button(
                "Download Processed Table as Excel",
                data=excel_data_specific,
                file_name="processed_specific_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_specific" # Unique key
            )

        # else: This case is handled by the initial 'if not df_initial.empty' block now
        # st.info("No data found for the predefined range. Please check your Excel file or the predefined range settings.")

    except Exception as e:
        st.error(f"An unexpected error occurred while processing the Excel file: {e}")
        st.exception(e) # Display full traceback for debugging
        st.info("Please ensure it's a valid Excel file with readable content and try again.")
else:
    st.info("Please upload your Excel file to begin processing.")
